import sqlite3
from pathlib import Path
from urllib.parse import urlparse


def _clean_text(value: str | None) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_url(uri: str) -> bool:
    parsed = urlparse(uri)
    return parsed.scheme in {"http", "https", "file", "ssh", "git"}


def validate_uri(uri: str, *, no_check: bool = False) -> tuple[bool, str]:
    uri = _clean_text(uri)
    if not uri:
        return False, "uri 不能为空"
    if is_url(uri):
        parsed = urlparse(uri)
        if parsed.scheme in {"http", "https", "ssh", "git"} and not parsed.netloc:
            return False, "URL 格式无效"
        return True, "url"
    if no_check:
        return True, "skip"
    if Path(uri).expanduser().exists():
        return True, "ok"
    return False, f"路径不存在：{uri}"


def upsert_resource(
    conn: sqlite3.Connection,
    *,
    project_id: int,
    resource_type: str,
    name: str,
    uri: str,
    is_dir: int = 1,
    tags: str = "",
    note: str = "",
    no_check: bool = False,
) -> int:
    ok, msg = validate_uri(uri, no_check=no_check)
    if not ok:
        raise RuntimeError(msg)

    cur = conn.execute(
        """
        INSERT INTO project_resources (project_id, type, name, uri, is_dir, tags, note)
        VALUES (?,?,?,?,?,?,?)
        ON CONFLICT(project_id, type, uri)
        DO UPDATE SET
            name=excluded.name,
            is_dir=excluded.is_dir,
            tags=excluded.tags,
            note=excluded.note,
            updated_at=datetime('now','localtime')
        """,
        (project_id, _clean_text(resource_type), _clean_text(name), _clean_text(uri), int(bool(is_dir)), _clean_text(tags) or None, _clean_text(note) or None),
    )
    row = conn.execute(
        "SELECT id FROM project_resources WHERE project_id=? AND type=? AND uri=?",
        (project_id, _clean_text(resource_type), _clean_text(uri)),
    ).fetchone()
    return int(row["id"]) if row else int(cur.lastrowid)


def list_resources(conn: sqlite3.Connection, project_id: int) -> list[sqlite3.Row]:
    return conn.execute(
        """
        SELECT id, type, name, uri, is_dir, tags, note, created_at, updated_at
        FROM project_resources
        WHERE project_id=?
        ORDER BY type, name, id
        """,
        (project_id,),
    ).fetchall()


def remove_resource(conn: sqlite3.Connection, project_id: int, resource_type: str, uri: str) -> int:
    cur = conn.execute(
        "DELETE FROM project_resources WHERE project_id=? AND type=? AND uri=?",
        (project_id, _clean_text(resource_type), _clean_text(uri)),
    )
    return int(cur.rowcount)


def check_resources(conn: sqlite3.Connection, project_id: int) -> list[dict]:
    rows = list_resources(conn, project_id)
    result: list[dict] = []
    for r in rows:
        ok, msg = validate_uri(r["uri"], no_check=False)
        result.append({
            "id": int(r["id"]),
            "type": r["type"],
            "name": r["name"],
            "uri": r["uri"],
            "ok": ok,
            "detail": msg,
        })
    return result


def import_resources_xlsx(
    conn: sqlite3.Connection,
    *,
    xlsx_path: Path,
    sheet: str,
    header_row: int,
    no_check: bool,
    auto_create_project: bool,
    get_project_id,
    create_project,
) -> tuple[int, int]:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)
    if sheet not in wb.sheetnames:
        raise RuntimeError(f"sheet 不存在：{sheet}")
    ws = wb[sheet]

    headers = [str(v).strip() if v is not None else "" for v in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
    col_map = {h: idx for idx, h in enumerate(headers) if h}
    required = ["project_code", "type", "name", "uri"]
    missing = [k for k in required if k not in col_map]
    if missing:
        raise RuntimeError(f"缺少列：{','.join(missing)}")

    ok = 0
    err = 0
    for row_idx, values in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if values is None or all(v in (None, "") for v in values):
            continue
        try:
            project_code = _clean_text(values[col_map["project_code"]])
            r_type = _clean_text(values[col_map["type"]])
            name = _clean_text(values[col_map["name"]])
            uri = _clean_text(values[col_map["uri"]])
            is_dir = int(values[col_map.get("is_dir", -1)] or 1) if "is_dir" in col_map else 1
            tags = _clean_text(values[col_map.get("tags", -1)] if "tags" in col_map else "")
            note = _clean_text(values[col_map.get("note", -1)] if "note" in col_map else "")
            if not project_code:
                raise RuntimeError("project_code 不能为空")
            try:
                project_id = int(get_project_id(project_code))
            except Exception:
                if not auto_create_project:
                    raise RuntimeError(f"项目不存在：{project_code}")
                create_project(project_code, project_code)
                project_id = int(get_project_id(project_code))
            upsert_resource(
                conn,
                project_id=project_id,
                resource_type=r_type,
                name=name,
                uri=uri,
                is_dir=is_dir,
                tags=tags,
                note=note,
                no_check=no_check,
            )
            ok += 1
        except Exception as exc:
            err += 1
            raise RuntimeError(f"第 {row_idx} 行导入失败：{exc}") from exc
    return ok, err
