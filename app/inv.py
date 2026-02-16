import argparse
import csv
import json
import os
import re
import sqlite3
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup

from app.project_resources import (
    check_resources,
    import_resources_xlsx,
    list_resources,
    remove_resource,
    upsert_resource,
)
from app.db import connect as db_connect


# ---------------------------
# 基础工具
# ---------------------------
def clean_text(s: str) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip()


def safe_filename(name: str) -> str:
    name = (name or "").strip()
    name = re.sub(r'[\\/:*?"<>|]+', "_", name)
    name = re.sub(r"\s+", "_", name)
    return name[:180] if len(name) > 180 else name


def normalize_url(url: str) -> str:
    p = urlparse(url)
    return f"{p.scheme}://{p.netloc}{p.path}"


def now_local_sql() -> str:
    return "datetime('now','localtime')"


def resolve_input_path(raw_path: str, cwd: Path) -> Path:
    """兼容 Windows 路径输入（如 G:/LabInventory/xxx）并映射到当前仓库。"""
    candidate = Path(raw_path)
    if candidate.exists():
        return candidate

    normalized = raw_path.replace('\\', '/')
    mapped = Path(normalized)
    if mapped.exists():
        return mapped

    m = re.match(r'^[A-Za-z]:/LabInventory/(.+)$', normalized)
    if m:
        local = cwd / m.group(1)
        if local.exists():
            return local

    local_by_name = cwd / Path(normalized).name
    if local_by_name.exists():
        return local_by_name

    return candidate


def resolve_output_path(raw_path: str, cwd: Path) -> Path:
    candidate = Path(raw_path)
    if os.name == 'nt':
        return candidate

    normalized = raw_path.replace('\\', '/')
    m = re.match(r'^[A-Za-z]:/LabInventory/(.+)$', normalized)
    if m:
        return cwd / m.group(1)
    return Path(normalized)


# ---------------------------
# 数据库初始化（幂等）
# ---------------------------
DDL = r"""
PRAGMA foreign_keys = ON;

CREATE TABLE IF NOT EXISTS parts (
  id            INTEGER PRIMARY KEY AUTOINCREMENT,
  mpn           TEXT NOT NULL,
  name          TEXT NOT NULL,
  category      TEXT NOT NULL,
  package       TEXT,
  params        TEXT,
  unit          TEXT NOT NULL DEFAULT 'pcs',
  url           TEXT,
  datasheet     TEXT,
  note          TEXT,
  created_at    TEXT NOT NULL DEFAULT (datetime('now','localtime'))
);
CREATE UNIQUE INDEX IF NOT EXISTS idx_parts_mpn ON parts(mpn);
CREATE INDEX IF NOT EXISTS idx_parts_search ON parts(name, category);

CREATE TABLE IF NOT EXISTS stock (
  id            INTEGER PRIMARY KEY AUTOINCREMENT,
  part_id       INTEGER NOT NULL,
  location      TEXT NOT NULL,
  qty           INTEGER NOT NULL DEFAULT 0,
  condition     TEXT NOT NULL DEFAULT 'new',
  updated_at    TEXT NOT NULL DEFAULT (datetime('now','localtime')),
  note          TEXT,
  FOREIGN KEY (part_id) REFERENCES parts(id) ON DELETE CASCADE
);
CREATE INDEX IF NOT EXISTS idx_stock_part ON stock(part_id);
CREATE INDEX IF NOT EXISTS idx_stock_loc  ON stock(location);

CREATE TABLE IF NOT EXISTS locations (
  location   TEXT PRIMARY KEY,
  note       TEXT
);

CREATE TABLE IF NOT EXISTS projects (
  id          INTEGER PRIMARY KEY AUTOINCREMENT,
  code        TEXT NOT NULL UNIQUE,
  name        TEXT NOT NULL,
  owner       TEXT,
  status      TEXT NOT NULL DEFAULT 'active',
  note        TEXT,
  created_at  TEXT NOT NULL DEFAULT (datetime('now','localtime'))
);

CREATE TABLE IF NOT EXISTS project_bom (
  id          INTEGER PRIMARY KEY AUTOINCREMENT,
  project_id  INTEGER NOT NULL,
  part_id     INTEGER NOT NULL,
  req_qty     INTEGER NOT NULL DEFAULT 1,
  priority    INTEGER NOT NULL DEFAULT 2,
  note        TEXT,
  UNIQUE(project_id, part_id),
  FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE,
  FOREIGN KEY (part_id) REFERENCES parts(id) ON DELETE CASCADE
);
CREATE INDEX IF NOT EXISTS idx_bom_project ON project_bom(project_id);
CREATE INDEX IF NOT EXISTS idx_bom_part    ON project_bom(part_id);

CREATE TABLE IF NOT EXISTS project_alloc (
  id          INTEGER PRIMARY KEY AUTOINCREMENT,
  project_id  INTEGER NOT NULL,
  part_id     INTEGER NOT NULL,
  location    TEXT,
  alloc_qty   INTEGER NOT NULL DEFAULT 0,
  status      TEXT NOT NULL DEFAULT 'reserved',
  note        TEXT,
  updated_at  TEXT NOT NULL DEFAULT (datetime('now','localtime')),
  FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE,
  FOREIGN KEY (part_id) REFERENCES parts(id) ON DELETE CASCADE
);
CREATE INDEX IF NOT EXISTS idx_alloc_project ON project_alloc(project_id);
CREATE INDEX IF NOT EXISTS idx_alloc_part    ON project_alloc(part_id);
CREATE INDEX IF NOT EXISTS idx_alloc_loc     ON project_alloc(location);

CREATE TABLE IF NOT EXISTS inv_doc (
  id            INTEGER PRIMARY KEY AUTOINCREMENT,
  doc_type      TEXT NOT NULL CHECK(doc_type IN ('IN','OUT','MOVE','ADJUST','CONSUME','RESERVE','RELEASE')),
  project_id    INTEGER,
  from_location TEXT,
  to_location   TEXT,
  ref           TEXT,
  operator      TEXT,
  note          TEXT,
  alloc_id      INTEGER,
  created_at    TEXT NOT NULL DEFAULT (datetime('now','localtime')),
  FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE SET NULL,
  FOREIGN KEY (alloc_id) REFERENCES project_alloc(id) ON DELETE SET NULL
);
CREATE INDEX IF NOT EXISTS idx_inv_doc_project ON inv_doc(project_id);
CREATE INDEX IF NOT EXISTS idx_inv_doc_created ON inv_doc(created_at);
CREATE INDEX IF NOT EXISTS idx_inv_doc_type    ON inv_doc(doc_type);

CREATE TABLE IF NOT EXISTS inv_line (
  id            INTEGER PRIMARY KEY AUTOINCREMENT,
  doc_id        INTEGER NOT NULL,
  part_id       INTEGER NOT NULL,
  qty           INTEGER NOT NULL CHECK(qty > 0),
  unit_cost     REAL,
  note          TEXT,
  FOREIGN KEY (doc_id) REFERENCES inv_doc(id) ON DELETE CASCADE,
  FOREIGN KEY (part_id) REFERENCES parts(id) ON DELETE RESTRICT
);
CREATE INDEX IF NOT EXISTS idx_inv_line_doc  ON inv_line(doc_id);
CREATE INDEX IF NOT EXISTS idx_inv_line_part ON inv_line(part_id);

-- 位置合法性（location 不为空时必须存在）
CREATE TRIGGER IF NOT EXISTS trg_alloc_location_check
BEFORE INSERT ON project_alloc
WHEN NEW.location IS NOT NULL AND NEW.location <> ''
BEGIN
  SELECT
    CASE
      WHEN (SELECT COUNT(1) FROM locations WHERE location = NEW.location) = 0
      THEN RAISE(ABORT, 'location 不存在于 locations 表')
    END;
END;

CREATE TRIGGER IF NOT EXISTS trg_alloc_location_check_u
BEFORE UPDATE OF location ON project_alloc
WHEN NEW.location IS NOT NULL AND NEW.location <> ''
BEGIN
  SELECT
    CASE
      WHEN (SELECT COUNT(1) FROM locations WHERE location = NEW.location) = 0
      THEN RAISE(ABORT, 'location 不存在于 locations 表')
    END;
END;

-- 强约束：禁止超预留（全局 + 库位）
DROP TRIGGER IF EXISTS trg_alloc_no_overreserve_ins;
DROP TRIGGER IF EXISTS trg_alloc_no_overreserve_upd;

CREATE TRIGGER trg_alloc_no_overreserve_ins
BEFORE INSERT ON project_alloc
WHEN NEW.alloc_qty > 0
 AND NEW.status IN ('reserved','consumed')
BEGIN
  SELECT
    CASE
      WHEN (
        IFNULL((SELECT SUM(qty) FROM stock WHERE part_id = NEW.part_id), 0)
        -
        IFNULL((SELECT SUM(alloc_qty) FROM project_alloc
                WHERE part_id = NEW.part_id
                  AND status IN ('reserved','consumed')
                  AND alloc_qty > 0), 0)
      ) < NEW.alloc_qty
      THEN RAISE(ABORT, '超预留：全局可用库存不足')
    END;

  SELECT
    CASE
      WHEN (NEW.location IS NOT NULL AND NEW.location <> '')
       AND (
        IFNULL((SELECT SUM(qty) FROM stock
                WHERE part_id = NEW.part_id AND location = NEW.location), 0)
        -
        IFNULL((SELECT SUM(alloc_qty) FROM project_alloc
                WHERE part_id = NEW.part_id
                  AND location = NEW.location
                  AND status IN ('reserved','consumed')
                  AND alloc_qty > 0), 0)
      ) < NEW.alloc_qty
      THEN RAISE(ABORT, '超预留：该库位可用库存不足')
    END;
END;

CREATE TRIGGER trg_alloc_no_overreserve_upd
BEFORE UPDATE OF part_id, location, alloc_qty, status ON project_alloc
WHEN NEW.alloc_qty > 0
 AND NEW.status IN ('reserved','consumed')
BEGIN
  SELECT
    CASE
      WHEN (
        IFNULL((SELECT SUM(qty) FROM stock WHERE part_id = NEW.part_id), 0)
        -
        IFNULL((SELECT SUM(alloc_qty) FROM project_alloc
                WHERE part_id = NEW.part_id
                  AND status IN ('reserved','consumed')
                  AND alloc_qty > 0
                  AND id <> OLD.id), 0)
      ) < NEW.alloc_qty
      THEN RAISE(ABORT, '超预留：全局可用库存不足（更新被阻止）')
    END;

  SELECT
    CASE
      WHEN (NEW.location IS NOT NULL AND NEW.location <> '')
       AND (
        IFNULL((SELECT SUM(qty) FROM stock
                WHERE part_id = NEW.part_id AND location = NEW.location), 0)
        -
        IFNULL((SELECT SUM(alloc_qty) FROM project_alloc
                WHERE part_id = NEW.part_id
                  AND location = NEW.location
                  AND status IN ('reserved','consumed')
                  AND alloc_qty > 0
                  AND id <> OLD.id), 0)
      ) < NEW.alloc_qty
      THEN RAISE(ABORT, '超预留：该库位可用库存不足（更新被阻止）')
    END;
END;

-- 视图：项目物料状态
CREATE VIEW IF NOT EXISTS v_project_material_status AS
WITH
stock_sum AS (
  SELECT part_id, SUM(qty) AS total_stock
  FROM stock
  GROUP BY part_id
),
alloc_sum_all AS (
  SELECT part_id, SUM(CASE WHEN status IN ('reserved','consumed') AND alloc_qty>0 THEN alloc_qty ELSE 0 END) AS reserved_qty
  FROM project_alloc
  GROUP BY part_id
),
alloc_sum_proj AS (
  SELECT project_id, part_id, SUM(CASE WHEN status IN ('reserved','consumed') AND alloc_qty>0 THEN alloc_qty ELSE 0 END) AS reserved_for_project
  FROM project_alloc
  GROUP BY project_id, part_id
)
SELECT
  pr.code AS project_code,
  pr.name AS project_name,
  p.category,
  p.mpn,
  p.name AS part_desc,
  p.package,
  p.params,
  b.req_qty,
  IFNULL(ss.total_stock, 0) AS total_stock,
  IFNULL(aa.reserved_qty, 0) AS reserved_qty_all_projects,
  (IFNULL(ss.total_stock, 0) - IFNULL(aa.reserved_qty, 0)) AS available_stock,
  IFNULL(ap.reserved_for_project, 0) AS reserved_for_project,
  (b.req_qty - IFNULL(ap.reserved_for_project, 0)) AS remaining_to_reserve,
  CASE
    WHEN (IFNULL(ss.total_stock, 0) - IFNULL(aa.reserved_qty, 0)) >= (b.req_qty - IFNULL(ap.reserved_for_project, 0))
      THEN 0
    ELSE (b.req_qty - IFNULL(ap.reserved_for_project, 0)) - (IFNULL(ss.total_stock, 0) - IFNULL(aa.reserved_qty, 0))
  END AS shortage_if_reserve_now
FROM project_bom b
JOIN projects pr ON pr.id = b.project_id
JOIN parts p     ON p.id  = b.part_id
LEFT JOIN stock_sum ss      ON ss.part_id = p.id
LEFT JOIN alloc_sum_all aa  ON aa.part_id = p.id
LEFT JOIN alloc_sum_proj ap ON ap.project_id = pr.id AND ap.part_id = p.id;

-- 视图：预留明细
CREATE VIEW IF NOT EXISTS v_project_alloc_detail AS
SELECT
  pr.code AS project_code,
  pr.name AS project_name,
  p.mpn,
  p.name AS part_desc,
  a.location,
  a.alloc_qty,
  a.status,
  a.updated_at,
  a.note
FROM project_alloc a
JOIN projects pr ON pr.id = a.project_id
JOIN parts p     ON p.id  = a.part_id;

CREATE TABLE IF NOT EXISTS project_resources (
  id          INTEGER PRIMARY KEY AUTOINCREMENT,
  project_id  INTEGER NOT NULL,
  type        TEXT    NOT NULL,
  name        TEXT    NOT NULL,
  uri         TEXT    NOT NULL,
  is_dir      INTEGER NOT NULL DEFAULT 1,
  note        TEXT,
  tags        TEXT,
  created_at  TEXT    NOT NULL DEFAULT (datetime('now','localtime')),
  updated_at  TEXT    NOT NULL DEFAULT (datetime('now','localtime')),
  FOREIGN KEY(project_id) REFERENCES projects(id) ON DELETE CASCADE
);
CREATE INDEX IF NOT EXISTS idx_project_resources_project ON project_resources(project_id);
CREATE INDEX IF NOT EXISTS idx_project_resources_type    ON project_resources(type);
CREATE UNIQUE INDEX IF NOT EXISTS uq_project_resources_unique
  ON project_resources(project_id, type, uri);

CREATE TRIGGER IF NOT EXISTS trg_project_resources_updated
AFTER UPDATE ON project_resources
FOR EACH ROW
BEGIN
  UPDATE project_resources
  SET updated_at = datetime('now','localtime')
  WHERE id = NEW.id;
END;

CREATE VIEW IF NOT EXISTS v_project_resources AS
SELECT
  p.code  AS project_code,
  p.name  AS project_name,
  r.type,
  r.name  AS resource_name,
  r.uri,
  r.is_dir,
  r.tags,
  r.note,
  r.updated_at
FROM project_resources r
JOIN projects p ON p.id = r.project_id;

CREATE VIEW IF NOT EXISTS v_project_overview AS
SELECT
  p.code AS project_code,
  p.name AS project_name,
  p.status,
  (SELECT COUNT(*) FROM project_bom b WHERE b.project_id = p.id) AS bom_lines,
  (SELECT COALESCE(SUM(b.req_qty),0) FROM project_bom b WHERE b.project_id = p.id) AS bom_total_req,
  (SELECT COUNT(*) FROM project_alloc a WHERE a.project_id = p.id) AS alloc_lines,
  (SELECT COALESCE(SUM(a.alloc_qty),0) FROM project_alloc a WHERE a.project_id = p.id) AS alloc_total_reserved,
  (SELECT COUNT(*) FROM project_resources r WHERE r.project_id = p.id) AS resource_count,
  p.created_at
FROM projects p;
"""

MIGRATION_DDL = r"""
CREATE TABLE IF NOT EXISTS inventory_txn (
  id            INTEGER PRIMARY KEY AUTOINCREMENT,
  txn_type      TEXT NOT NULL CHECK(txn_type IN ('IN','OUT','ADJUST')),
  project_id    INTEGER,
  ref           TEXT,
  note          TEXT,
  operator      TEXT,
  created_at    TEXT NOT NULL DEFAULT (datetime('now','localtime')),
  FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE SET NULL
);
CREATE INDEX IF NOT EXISTS idx_inventory_txn_type    ON inventory_txn(txn_type);
CREATE INDEX IF NOT EXISTS idx_inventory_txn_created ON inventory_txn(created_at);
CREATE INDEX IF NOT EXISTS idx_inventory_txn_project ON inventory_txn(project_id);

CREATE TABLE IF NOT EXISTS inventory_txn_line (
  id            INTEGER PRIMARY KEY AUTOINCREMENT,
  txn_id        INTEGER NOT NULL,
  part_id       INTEGER NOT NULL,
  mpn_snapshot  TEXT NOT NULL,
  location      TEXT NOT NULL,
  qty_delta     INTEGER NOT NULL,
  condition     TEXT NOT NULL DEFAULT 'new',
  note          TEXT,
  created_at    TEXT NOT NULL DEFAULT (datetime('now','localtime')),
  FOREIGN KEY (txn_id) REFERENCES inventory_txn(id) ON DELETE CASCADE,
  FOREIGN KEY (part_id) REFERENCES parts(id) ON DELETE RESTRICT
);
CREATE INDEX IF NOT EXISTS idx_inventory_txn_line_txn      ON inventory_txn_line(txn_id);
CREATE INDEX IF NOT EXISTS idx_inventory_txn_line_part     ON inventory_txn_line(part_id);
CREATE INDEX IF NOT EXISTS idx_inventory_txn_line_location ON inventory_txn_line(location);
"""


def connect(db_path: Path) -> sqlite3.Connection:
    return db_connect(db_path)


def init_db(conn: sqlite3.Connection):
    conn.executescript(DDL)
    apply_migrations(conn)
    conn.commit()


def apply_migrations(conn: sqlite3.Connection):
    conn.executescript(MIGRATION_DDL)


# ---------------------------
# 立创抓取 + 数据手册下载
# ---------------------------
def find_value_by_label(soup: BeautifulSoup, label: str) -> str:
    text = soup.get_text("\n", strip=True)
    m = re.search(rf"{re.escape(label)}\s*\n([^\n]+)", text)
    if m:
        return clean_text(m.group(1))
    m = re.search(rf"{re.escape(label)}\s*[:：]?\s*([^\n]+)", text)
    if m:
        return clean_text(m.group(1))
    return ""


def parse_params_table(soup: BeautifulSoup) -> dict:
    text = soup.get_text("\n", strip=True)
    if "商品参数" not in text:
        return {}
    tail = text.split("商品参数", 1)[1]
    lines = [clean_text(x) for x in tail.split("\n") if clean_text(x)]

    start_idx = None
    for i in range(len(lines) - 1):
        if lines[i] == "属性" and lines[i + 1] == "参数值":
            start_idx = i + 2
            break
    if start_idx is None:
        return {}

    end_signals = {"相似推荐", "其他推荐", "客服", "反馈", "收起", "置顶"}
    pairs = {}
    i = start_idx
    while i + 1 < len(lines):
        if lines[i] in end_signals:
            break
        key = lines[i]
        val = lines[i + 1]
        if key and val and key not in {"属性", "参数值"} and key not in pairs:
            pairs[key] = val
        i += 2
    return pairs


def find_datasheet_url(soup: BeautifulSoup, base_url: str) -> str:
    candidates = []

    for a in soup.find_all("a", href=True):
        txt = clean_text(a.get_text(" ", strip=True))
        href = a["href"]
        full = urljoin(base_url, href)
        if re.search(r"(数据手册|Datasheet)", txt, re.I):
            candidates.append(full)

    for a in soup.find_all("a", href=True):
        href = a["href"]
        full = urljoin(base_url, href)
        if ".pdf" in full.lower():
            candidates.append(full)

    raw = str(soup)
    for m in re.finditer(r"https?://[^\s\"']+\.pdf", raw, flags=re.I):
        candidates.append(m.group(0))

    seen, uniq = set(), []
    for u in candidates:
        u = u.strip()
        if u and u not in seen:
            seen.add(u)
            uniq.append(u)

    for u in uniq:
        if u.lower().endswith(".pdf"):
            return u
    for u in uniq:
        if ".pdf" in u.lower():
            return u
    return ""


def download_pdf(session: requests.Session, pdf_url: str, out_path: Path) -> bool:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with session.get(pdf_url, stream=True, timeout=30) as r:
        r.raise_for_status()
        ctype = (r.headers.get("Content-Type") or "").lower()

        # content-type 异常时检查 PDF 头
        if ("pdf" not in ctype) and ("octet-stream" not in ctype):
            first = r.raw.read(5)
            if first != b"%PDF-":
                return False
            with open(out_path, "wb") as f:
                f.write(first)
                for chunk in r.iter_content(chunk_size=1024 * 128):
                    if chunk:
                        f.write(chunk)
            return True

        with open(out_path, "wb") as f:
            for chunk in r.iter_content(chunk_size=1024 * 128):
                if chunk:
                    f.write(chunk)

    return out_path.exists() and out_path.stat().st_size > 1024


@dataclass
class LcscItem:
    mpn: str
    desc: str
    category: str
    package: str
    brand: str
    lcsc_code: str
    params_text: str
    page_url: str
    datasheet_local: str
    note: str


def lcsc_fetch_and_parse(url: str, datasheets_dir: Path) -> LcscItem:
    session = requests.Session()
    session.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/120.0 Safari/537.36"
        ),
        "Accept-Language": "zh-CN,zh;q=0.9",
    })

    page_url = normalize_url(url)
    html = session.get(page_url, timeout=20)
    html.raise_for_status()
    html.encoding = html.apparent_encoding or "utf-8"

    soup = BeautifulSoup(html.text, "lxml")

    mpn = find_value_by_label(soup, "商品型号")
    if not mpn:
        t = soup.get_text("\n", strip=True)
        m = re.search(r"商品型号\s*[:：]?\s*([A-Za-z0-9\-_.]+)", t)
        mpn = m.group(1) if m else ""
    if not mpn:
        raise RuntimeError("解析失败：未找到商品型号")

    desc = find_value_by_label(soup, "描述") or mpn
    category = find_value_by_label(soup, "商品目录") or "未分类"
    package = find_value_by_label(soup, "商品封装")
    brand = find_value_by_label(soup, "品牌名称")
    lcsc_code = find_value_by_label(soup, "商品编号")

    params = parse_params_table(soup)
    params_text = "; ".join([f"{k}={v}" for k, v in params.items()]) if params else ""

    # datasheet：默认商品页，下载成功则本地路径
    datasheet_value = page_url
    pdf_url = find_datasheet_url(soup, page_url)

    note_parts = []
    if brand:
        note_parts.append(f"品牌={brand}")
    if lcsc_code:
        note_parts.append(f"LCSC={lcsc_code}")
    note_parts.append(f"URL={page_url}")

    datasheet_local = ""
    if pdf_url:
        base = safe_filename(mpn)
        if lcsc_code:
            base += f"__{safe_filename(lcsc_code)}"
        out_pdf = datasheets_dir / f"{base}.pdf"
        try:
            if download_pdf(session, pdf_url, out_pdf):
                datasheet_value = str(out_pdf)
                datasheet_local = str(out_pdf)
                note_parts.append(f"DatasheetPDF={pdf_url}")
            else:
                note_parts.append(f"DatasheetPDF下载失败={pdf_url}")
        except Exception as e:
            note_parts.append(f"DatasheetPDF异常={pdf_url} ({type(e).__name__})")
    else:
        note_parts.append("DatasheetPDF未找到")

    note = " | ".join(note_parts)

    return LcscItem(
        mpn=mpn,
        desc=desc,
        category=category,
        package=package,
        brand=brand,
        lcsc_code=lcsc_code,
        params_text=params_text,
        page_url=page_url,
        datasheet_local=datasheet_local,
        note=note,
    )


# ---------------------------
# DB 操作函数（全部通过脚本）
# ---------------------------
def get_project_id(conn, code: str) -> int:
    r = conn.execute("SELECT id FROM projects WHERE code=?", (code,)).fetchone()
    if not r:
        raise RuntimeError(f"项目不存在：{code}")
    return int(r["id"])


def get_part_id_by_mpn(conn, mpn: str) -> int:
    r = conn.execute("SELECT id FROM parts WHERE mpn=?", (mpn,)).fetchone()
    if not r:
        raise RuntimeError(f"物料不存在：{mpn}")
    return int(r["id"])


def get_project_id_optional(conn, code: str = "") -> int | None:
    code = clean_text(code)
    if not code:
        return None
    return get_project_id(conn, code)


def assert_location_exists(conn, location: str):
    if conn.execute("SELECT 1 FROM locations WHERE location=?", (location,)).fetchone() is None:
        raise RuntimeError(f"库位不存在（locations 表里没有）：{location}")


def _tx_begin(conn, tx_name: str = "inv_tx"):
    conn.execute(f"SAVEPOINT {tx_name};")


def _tx_commit(conn, tx_name: str = "inv_tx"):
    conn.execute(f"RELEASE SAVEPOINT {tx_name};")


def _tx_rollback(conn, tx_name: str = "inv_tx"):
    conn.execute(f"ROLLBACK TO SAVEPOINT {tx_name};")
    conn.execute(f"RELEASE SAVEPOINT {tx_name};")


def write_ledger(
    conn,
    *,
    doc_type: str,
    part_id: int,
    qty: int,
    project_id: int | None = None,
    from_location: str | None = None,
    to_location: str | None = None,
    ref: str = "",
    operator: str = "",
    note: str = "",
    alloc_id: int | None = None,
    unit_cost: float | None = None,
    line_note: str = "",
) -> int:
    if qty <= 0:
        raise RuntimeError("ledger qty 必须为正整数")
    cur = conn.execute(
        """
        INSERT INTO inv_doc (doc_type, project_id, from_location, to_location, ref, operator, note, alloc_id)
        VALUES (?,?,?,?,?,?,?,?)
        """,
        (doc_type, project_id, from_location, to_location, ref, operator, note, alloc_id),
    )
    doc_id = int(cur.lastrowid)
    conn.execute(
        "INSERT INTO inv_line (doc_id, part_id, qty, unit_cost, note) VALUES (?,?,?,?,?)",
        (doc_id, part_id, qty, unit_cost, line_note or None),
    )
    return doc_id


def create_txn(conn, txn_type: str, project_code: str | None = None, ref: str = "", note: str = "", operator: str = "") -> int:
    txn_type = clean_text(txn_type).upper()
    if txn_type not in {"IN", "OUT", "ADJUST"}:
        raise RuntimeError(f"不支持的 txn_type：{txn_type}")
    project_id = get_project_id_optional(conn, project_code or "")
    cur = conn.execute(
        "INSERT INTO inventory_txn (txn_type, project_id, ref, note, operator) VALUES (?,?,?,?,?)",
        (txn_type, project_id, ref or None, note or None, operator or None),
    )
    return int(cur.lastrowid)


def add_txn_line(
    conn,
    txn_id: int,
    mpn: str,
    location: str,
    qty_delta: int,
    condition: str = "new",
    note: str = "",
) -> int:
    if qty_delta == 0:
        raise RuntimeError("qty_delta 不能为 0")
    assert_location_exists(conn, location)
    part_id = get_part_id_by_mpn(conn, mpn)
    cur = conn.execute(
        """
        INSERT INTO inventory_txn_line (txn_id, part_id, mpn_snapshot, location, qty_delta, condition, note)
        VALUES (?,?,?,?,?,?,?)
        """,
        (txn_id, part_id, mpn, location, qty_delta, condition, note or None),
    )
    return int(cur.lastrowid)


def apply_stock_delta(conn, *, part_id: int, location: str, qty_delta: int, condition: str = "new", note: str = ""):
    if qty_delta == 0:
        raise RuntimeError("库存变化量不能为 0")
    row = conn.execute("SELECT id, qty FROM stock WHERE part_id=? AND location=?", (part_id, location)).fetchone()
    current = int(row["qty"]) if row else 0
    nxt = current + qty_delta
    if nxt < 0:
        raise RuntimeError(f"库存不足：part_id={part_id} location={location} stock={current} delta={qty_delta}")
    if row:
        conn.execute(
            "UPDATE stock SET qty=?, updated_at=datetime('now','localtime'), condition=?, note=? WHERE id=?",
            (nxt, condition, note, int(row["id"])),
        )
    else:
        conn.execute(
            "INSERT INTO stock (part_id, location, qty, condition, note) VALUES (?,?,?,?,?)",
            (part_id, location, qty_delta, condition, note),
        )


def upsert_part(conn, mpn: str, name: str, category: str, package: str, params: str, url: str, datasheet: str, note: str) -> int:
    row = conn.execute("SELECT id FROM parts WHERE mpn=?", (mpn,)).fetchone()
    if row:
        pid = int(row["id"])
        conn.execute(
            """
            UPDATE parts
            SET name=COALESCE(NULLIF(?,''), name),
                category=COALESCE(NULLIF(?,''), category),
                package=COALESCE(NULLIF(?,''), package),
                params=COALESCE(NULLIF(?,''), params),
                url=COALESCE(NULLIF(?,''), url),
                datasheet=COALESCE(NULLIF(?,''), datasheet),
                note=COALESCE(NULLIF(?,''), note)
            WHERE id=?
            """,
            (name, category, package, params, url, datasheet, note, pid),
        )
        return pid
    cur = conn.execute(
        "INSERT INTO parts (mpn, name, category, package, params, url, datasheet, note) VALUES (?,?,?,?,?,?,?,?)",
        (mpn, name, category, package, params, url, datasheet, note),
    )
    return int(cur.lastrowid)


def stock_in(
    conn,
    mpn: str,
    location: str,
    qty: int,
    condition: str = "new",
    note: str = "",
    project_code: str = "",
    ref: str = "",
    operator: str = "",
):
    if qty <= 0:
        raise RuntimeError("入库数量必须为正整数")
    assert_location_exists(conn, location)
    part_id = get_part_id_by_mpn(conn, mpn)
    project_id = get_project_id_optional(conn, project_code)
    _tx_begin(conn)
    try:
        apply_stock_delta(conn, part_id=part_id, location=location, qty_delta=qty, condition=condition, note=note)
        txn_id = create_txn(conn, "IN", project_code, ref=ref, note=note, operator=operator)
        add_txn_line(conn, txn_id, mpn, location, qty, condition=condition, note=note)
        write_ledger(
            conn,
            doc_type="IN",
            part_id=part_id,
            qty=qty,
            project_id=project_id,
            to_location=location,
            ref=ref,
            operator=operator,
            note=note,
        )
        _tx_commit(conn)
    except Exception:
        _tx_rollback(conn)
        raise


def add_stock(conn, mpn: str, location: str, qty: int, condition: str = "new", note: str = ""):
    # 兼容旧接口：默认作为入库处理，并同步写入 ledger(IN)
    stock_in(conn, mpn, location, qty, condition=condition, note=note)


def stock_out(conn, mpn: str, location: str, qty: int, project_code: str = "", ref: str = "", note: str = "", operator: str = ""):
    if qty <= 0:
        raise RuntimeError("出库数量必须为正整数")
    assert_location_exists(conn, location)
    part_id = get_part_id_by_mpn(conn, mpn)
    project_id = get_project_id_optional(conn, project_code)
    _tx_begin(conn)
    try:
        apply_stock_delta(conn, part_id=part_id, location=location, qty_delta=-qty, note=note)
        txn_id = create_txn(conn, "OUT", project_code, ref=ref, note=note, operator=operator)
        add_txn_line(conn, txn_id, mpn, location, -qty, note=note)
        write_ledger(
            conn,
            doc_type="OUT",
            part_id=part_id,
            qty=qty,
            project_id=project_id,
            from_location=location,
            ref=ref,
            operator=operator,
            note=note,
        )
        _tx_commit(conn)
    except Exception:
        _tx_rollback(conn)
        raise


def stock_move(conn, mpn: str, from_location: str, to_location: str, qty: int, note: str = "", operator: str = ""):
    if qty <= 0:
        raise RuntimeError("移库数量必须为正整数")
    if from_location == to_location:
        raise RuntimeError("from/to 不能相同")
    assert_location_exists(conn, from_location)
    assert_location_exists(conn, to_location)
    part_id = get_part_id_by_mpn(conn, mpn)
    src = conn.execute("SELECT id, qty FROM stock WHERE part_id=? AND location=?", (part_id, from_location)).fetchone()
    if not src:
        raise RuntimeError(f"源库位无库存记录：part_id={part_id}, location={from_location}")
    if int(src["qty"]) < qty:
        raise RuntimeError(f"移库失败：源库位库存不足（stock={int(src['qty'])} < move={qty}）")
    _tx_begin(conn)
    try:
        conn.execute(
            "UPDATE stock SET qty=qty-?, updated_at=datetime('now','localtime') WHERE id=?",
            (qty, int(src["id"])),
        )
        dst = conn.execute("SELECT id, qty, condition FROM stock WHERE part_id=? AND location=?", (part_id, to_location)).fetchone()
        if dst:
            conn.execute(
                "UPDATE stock SET qty=qty+?, updated_at=datetime('now','localtime'), note=? WHERE id=?",
                (qty, note, int(dst["id"])),
            )
        else:
            conn.execute(
                "INSERT INTO stock (part_id, location, qty, condition, note) VALUES (?,?,?,?,?)",
                (part_id, to_location, qty, "new", note),
            )
        write_ledger(
            conn,
            doc_type="MOVE",
            part_id=part_id,
            qty=qty,
            from_location=from_location,
            to_location=to_location,
            operator=operator,
            note=note,
        )
        _tx_commit(conn)
    except Exception:
        _tx_rollback(conn)
        raise


def stock_adjust(
    conn,
    mpn: str,
    location: str,
    add_qty: int = 0,
    sub_qty: int = 0,
    note: str = "",
    ref: str = "",
    operator: str = "",
):
    if bool(add_qty > 0) == bool(sub_qty > 0):
        raise RuntimeError("stock-adjust 必须且只能指定 --add 或 --sub")
    if not clean_text(note):
        raise RuntimeError("stock-adjust 必须提供 --note")
    assert_location_exists(conn, location)
    part_id = get_part_id_by_mpn(conn, mpn)
    _tx_begin(conn)
    try:
        if add_qty > 0:
            apply_stock_delta(conn, part_id=part_id, location=location, qty_delta=add_qty, note=note)
            qty = add_qty
            qty_delta = add_qty
        else:
            apply_stock_delta(conn, part_id=part_id, location=location, qty_delta=-sub_qty, note=note)
            qty = sub_qty
            qty_delta = -sub_qty
        txn_id = create_txn(conn, "ADJUST", None, ref=ref, note=note, operator=operator)
        add_txn_line(conn, txn_id, mpn, location, qty_delta, note=("add" if add_qty > 0 else "sub"))
        write_ledger(
            conn,
            doc_type="ADJUST",
            part_id=part_id,
            qty=qty,
            from_location=location if sub_qty > 0 else None,
            to_location=location if add_qty > 0 else None,
            ref=ref,
            operator=operator,
            note=note,
            line_note=("add" if add_qty > 0 else "sub"),
        )
        _tx_commit(conn)
    except Exception:
        _tx_rollback(conn)
        raise


def create_project(conn, code: str, name: str, owner: str = "", note: str = ""):
    conn.execute(
        "INSERT INTO projects (code, name, owner, note) VALUES (?,?,?,?)",
        (code, name, owner or None, note or None),
    )


def add_project(conn, code: str, name: str, owner: str = "", note: str = "") -> tuple[int, bool]:
    row = conn.execute("SELECT id FROM projects WHERE code=?", (code,)).fetchone()
    if row:
        conn.execute(
            """
            UPDATE projects
            SET name=COALESCE(NULLIF(?,''), name),
                owner=COALESCE(NULLIF(?,''), owner),
                note=COALESCE(NULLIF(?,''), note)
            WHERE code=?
            """,
            (name, owner, note, code),
        )
        return int(row["id"]), False
    create_project(conn, code, name, owner, note)
    new_id = int(conn.execute("SELECT id FROM projects WHERE code=?", (code,)).fetchone()["id"])
    return new_id, True


def show_project_resources(conn, project_code: str):
    project_id = get_project_id(conn, project_code)
    rows = list_resources(conn, project_id)
    if not rows:
        print("没有项目资源记录")
        return
    headers = ["id", "type", "name", "uri", "is_dir", "tags", "note", "updated_at"]
    print("\t".join(headers))
    for r in rows:
        print("\t".join(str(r[h] if r[h] is not None else "") for h in headers))


def show_project_overview(conn, project_code: str = ""):
    if clean_text(project_code):
        rows = conn.execute("SELECT * FROM v_project_overview WHERE project_code=?", (project_code,)).fetchall()
    else:
        rows = conn.execute("SELECT * FROM v_project_overview ORDER BY project_code").fetchall()
    if not rows:
        print("没有项目记录")
        return
    headers = ["project_code", "project_name", "status", "bom_lines", "bom_total_req", "alloc_lines", "alloc_total_reserved", "resource_count", "created_at"]
    print("\t".join(headers))
    for r in rows:
        print("\t".join(str(r[h]) for h in headers))


def check_project_resources(conn, project_code: str):
    project_id = get_project_id(conn, project_code)
    checks = check_resources(conn, project_id)
    if not checks:
        print("没有项目资源记录")
        return
    headers = ["id", "type", "name", "ok", "detail", "uri"]
    print("\t".join(headers))
    for item in checks:
        print("\t".join(str(item[h]) for h in headers))


def set_bom(conn, project_code: str, mpn: str, req_qty: int, priority: int = 2, note: str = ""):
    pid = get_project_id(conn, project_code)
    part_id = get_part_id_by_mpn(conn, mpn)
    # upsert bom
    row = conn.execute("SELECT id FROM project_bom WHERE project_id=? AND part_id=?", (pid, part_id)).fetchone()
    if row:
        conn.execute(
            "UPDATE project_bom SET req_qty=?, priority=?, note=? WHERE id=?",
            (req_qty, priority, note, int(row["id"])),
        )
    else:
        conn.execute(
            "INSERT INTO project_bom (project_id, part_id, req_qty, priority, note) VALUES (?,?,?,?,?)",
            (pid, part_id, req_qty, priority, note),
        )


def reserve_loc(conn, project_code: str, mpn: str, location: str, qty: int, note: str = "") -> int:
    if qty <= 0:
        raise RuntimeError("预留数量必须为正整数")
    # location 合法性（触发器也会检查，这里提前报错更友好）
    if conn.execute("SELECT 1 FROM locations WHERE location=?", (location,)).fetchone() is None:
        raise RuntimeError(f"库位不存在：{location}")

    pid = get_project_id(conn, project_code)
    part_id = get_part_id_by_mpn(conn, mpn)

    cur = conn.execute(
        "INSERT INTO project_alloc (project_id, part_id, location, alloc_qty, status, note, updated_at) VALUES (?,?,?,?, 'reserved', ?, datetime('now','localtime'))",
        (pid, part_id, location, qty, note),
    )
    alloc_id = int(cur.lastrowid)
    write_ledger(
        conn,
        doc_type="RESERVE",
        part_id=part_id,
        qty=qty,
        project_id=pid,
        to_location=location,
        note=note,
        alloc_id=alloc_id,
    )
    return alloc_id


def release_alloc(conn, alloc_id: int, note_append: str = "释放"):
    row = conn.execute("SELECT project_id, part_id, location, alloc_qty, status FROM project_alloc WHERE id=?", (alloc_id,)).fetchone()
    if not row:
        raise RuntimeError(f"alloc_id 不存在：{alloc_id}")
    conn.execute(
        "UPDATE project_alloc SET status='released', updated_at=datetime('now','localtime'), note=COALESCE(note,'') || ? WHERE id=?",
        (f" | {note_append}", alloc_id),
    )
    write_ledger(
        conn,
        doc_type="RELEASE",
        part_id=int(row["part_id"]),
        qty=int(row["alloc_qty"]),
        project_id=int(row["project_id"]) if row["project_id"] is not None else None,
        from_location=clean_text(row["location"]),
        note=note_append,
        alloc_id=alloc_id,
    )


def consume_alloc(conn, alloc_id: int, note_append: str = "已消耗"):
    """
    消耗 = 将 alloc 标记为 consumed + 扣减 stock（同一 part + location）
    强约束已保证 alloc 不会超预留，但 stock 扣减还需要确保该库位 stock 行存在且足够。
    """
    a = conn.execute("SELECT project_id, part_id, location, alloc_qty, status FROM project_alloc WHERE id=?", (alloc_id,)).fetchone()
    if not a:
        raise RuntimeError(f"alloc_id 不存在：{alloc_id}")
    if a["status"] != "reserved":
        raise RuntimeError(f"只有 reserved 状态才能消耗，当前={a['status']}")

    part_id = int(a["part_id"])
    location = clean_text(a["location"])
    qty = int(a["alloc_qty"])
    if not location:
        raise RuntimeError("带库位预留要求 location 非空，当前记录没有 location")

    s = conn.execute("SELECT id, qty FROM stock WHERE part_id=? AND location=?", (part_id, location)).fetchone()
    if not s:
        raise RuntimeError(f"stock 中找不到该库位记录：part_id={part_id}, location={location}")
    if int(s["qty"]) < qty:
        raise RuntimeError(f"扣减失败：库位库存不足（stock={int(s['qty'])} < consume={qty}）")

    # 事务：要么都成功要么都失败
    _tx_begin(conn)
    try:
        conn.execute(
            "UPDATE project_alloc SET status='consumed', updated_at=datetime('now','localtime'), note=COALESCE(note,'') || ? WHERE id=?",
            (f" | {note_append}", alloc_id),
        )
        conn.execute(
            "UPDATE stock SET qty=qty-?, updated_at=datetime('now','localtime') WHERE id=?",
            (qty, int(s["id"])),
        )
        project_code = ""
        if a["project_id"] is not None:
            pr = conn.execute("SELECT code FROM projects WHERE id=?", (int(a["project_id"]),)).fetchone()
            project_code = clean_text(pr["code"]) if pr else ""
        mpn_row = conn.execute("SELECT mpn FROM parts WHERE id=?", (part_id,)).fetchone()
        txn_id = create_txn(conn, "OUT", project_code, note=note_append)
        add_txn_line(conn, txn_id, clean_text(mpn_row["mpn"]) if mpn_row else str(part_id), location, -qty, note=note_append)
        write_ledger(
            conn,
            doc_type="CONSUME",
            part_id=part_id,
            qty=qty,
            project_id=int(a["project_id"]) if a["project_id"] is not None else None,
            from_location=location,
            note=note_append,
            alloc_id=alloc_id,
        )
        _tx_commit(conn)
    except Exception:
        _tx_rollback(conn)
        raise


def show_ledger(conn, project_code: str = "", mpn: str = "", since: str = ""):
    sql = """
    SELECT d.created_at, d.doc_type, pr.code AS project_code, p.mpn,
           d.from_location, d.to_location, l.qty, d.ref, d.operator, d.note
    FROM inv_doc d
    JOIN inv_line l ON l.doc_id = d.id
    JOIN parts p ON p.id = l.part_id
    LEFT JOIN projects pr ON pr.id = d.project_id
    WHERE 1=1
    """
    params = []
    if clean_text(project_code):
        sql += " AND pr.code = ?"
        params.append(project_code)
    if clean_text(mpn):
        sql += " AND p.mpn = ?"
        params.append(mpn)
    if clean_text(since):
        sql += " AND date(d.created_at) >= date(?)"
        params.append(since)
    sql += " ORDER BY d.id DESC"
    rows = conn.execute(sql, tuple(params)).fetchall()
    if not rows:
        print("没有流水记录")
        return
    headers = ["created_at", "doc_type", "project_code", "mpn", "from_location", "to_location", "qty", "ref", "operator", "note"]
    print("\t".join(headers))
    for r in rows:
        print("\t".join(str(r[h]) if r[h] is not None else "" for h in headers))


def export_schema_sql(conn) -> str:
    rows = conn.execute(
        """
        SELECT type, name, sql
        FROM sqlite_master
        WHERE sql IS NOT NULL AND name NOT LIKE 'sqlite_%'
        ORDER BY CASE type WHEN 'table' THEN 1 WHEN 'index' THEN 2 WHEN 'view' THEN 3 WHEN 'trigger' THEN 4 ELSE 9 END, name
        """
    ).fetchall()
    return "\n\n".join(f"-- {r['type']}: {r['name']}\n{r['sql']};" for r in rows)


def export_schema_md(conn) -> str:
    lines = ["# Database Schema", ""]
    for t in ("table", "view", "index", "trigger"):
        rows = conn.execute(
            "SELECT name, sql FROM sqlite_master WHERE type=? AND name NOT LIKE 'sqlite_%' ORDER BY name",
            (t,),
        ).fetchall()
        lines.append(f"## {t}s")
        lines.append("")
        for r in rows:
            lines.append(f"### {r['name']}")
            if t == "table":
                cols = conn.execute(f"PRAGMA table_info('{r['name']}')").fetchall()
                lines.append("| cid | name | type | notnull | dflt | pk |")
                lines.append("|---:|---|---|---:|---|---:|")
                for c in cols:
                    lines.append(f"| {c['cid']} | {c['name']} | {c['type']} | {c['notnull']} | {c['dflt_value'] or ''} | {c['pk']} |")
            lines.append("```sql")
            lines.append((r["sql"] or "").strip())
            lines.append("```")
            lines.append("")
    return "\n".join(lines)


def schema_export(conn, fmt: str = "sql", out_path: Path | None = None):
    content = export_schema_sql(conn) if fmt == "sql" else export_schema_md(conn)
    if out_path:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(content, encoding="utf-8")
        return
    print(content)


def _load_openpyxl():
    try:
        from openpyxl import Workbook, load_workbook
        return Workbook, load_workbook
    except Exception as e:
        raise RuntimeError("缺少依赖 openpyxl，请先安装：python -m pip install openpyxl") from e


def txn_export_xlsx_template(out_path: Path):
    Workbook, _ = _load_openpyxl()
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws.append(["txn_type", "project_code", "mpn", "location", "qty", "condition", "note", "ref", "operator"])
    ws.append(["IN", "", "SN74LVC1G08DBVR", "C409-G01-S01-P01", 10, "new", "样例入库", "BATCH-001", "alice"])
    ws.append(["OUT", "PJ-001", "SN74LVC1G08DBVR", "C409-G01-S01-P01", 2, "new", "样例出库", "BATCH-001", "alice"])

    ws_in = wb.create_sheet("StockIn")
    ws_in.append(["project_code", "mpn", "location", "qty", "condition", "note", "ref", "operator"])
    ws_in.append(["", "SN74LVC1G08DBVR", "C409-G01-S01-P01", 10, "new", "批量入库样例", "BATCH-001", "alice"])

    ws_out = wb.create_sheet("StockOut")
    ws_out.append(["project_code", "mpn", "location", "qty", "note", "ref", "operator"])
    ws_out.append(["PJ-001", "SN74LVC1G08DBVR", "C409-G01-S01-P01", 2, "批量出库样例", "BATCH-001", "alice"])
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _iter_txn_rows_from_workbook(wb, mode: str):
    rows = []
    has_transactions = "Transactions" in wb.sheetnames
    has_stock_io = ("StockIn" in wb.sheetnames) or ("StockOut" in wb.sheetnames)
    if mode == "transactions":
        if not has_transactions:
            raise RuntimeError("XLSX 缺少 Transactions sheet")
    elif mode == "stock-io":
        if not has_stock_io:
            raise RuntimeError("XLSX 缺少 StockIn/StockOut sheet")
    elif not has_transactions and not has_stock_io:
        raise RuntimeError("XLSX 缺少可识别的 sheet（Transactions 或 StockIn/StockOut）")

    if has_transactions and mode in ("auto", "transactions"):
        ws = wb["Transactions"]
        for idx, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            rows.append(("Transactions", idx, r))

    if has_stock_io and mode in ("auto", "stock-io"):
        if "StockIn" in wb.sheetnames:
            ws = wb["StockIn"]
            for idx, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                mapped = ("IN", r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7])
                rows.append(("StockIn", idx, mapped))
        if "StockOut" in wb.sheetnames:
            ws = wb["StockOut"]
            for idx, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                mapped = ("OUT", r[0], r[1], r[2], r[3], "new", r[4], r[5], r[6])
                rows.append(("StockOut", idx, mapped))
    return rows


def txn_import_xlsx(conn, xlsx_path: Path, partial: bool = False, error_out: Path | None = None, mode: str = "auto") -> tuple[int, int]:
    _, load_workbook = _load_openpyxl()
    wb = load_workbook(xlsx_path)
    rows = _iter_txn_rows_from_workbook(wb, mode=mode)
    errors = []
    ok = 0
    _tx_begin(conn, "xlsx_batch")
    try:
        for sheet_name, idx, r in rows:
            if r is None or all(v in (None, "") for v in r):
                continue
            txn_type = clean_text(r[0]).upper()
            project_code = clean_text(r[1])
            mpn = clean_text(r[2])
            location = clean_text(r[3])
            qty_raw = r[4]
            condition = clean_text(r[5]) or "new"
            note = clean_text(r[6])
            ref = clean_text(r[7])
            operator = clean_text(r[8])
            try:
                qty = int(qty_raw)
                if qty <= 0:
                    raise RuntimeError("qty 必须为正整数")
                if txn_type == "IN":
                    stock_in(conn, mpn, location, qty, condition=condition, note=note, project_code=project_code, ref=ref, operator=operator)
                elif txn_type == "OUT":
                    stock_out(conn, mpn, location, qty, project_code=project_code, ref=ref, note=note, operator=operator)
                elif txn_type == "ADJUST":
                    stock_adjust(conn, mpn, location, add_qty=qty, note=note or "xlsx adjust", ref=ref, operator=operator)
                else:
                    raise RuntimeError("txn_type 仅支持 IN/OUT/ADJUST")
                ok += 1
            except Exception as e:
                errors.append({"sheet": sheet_name, "row": idx, "error": str(e), "values": ["" if v is None else str(v) for v in r]})
                if not partial:
                    raise
        _tx_commit(conn, "xlsx_batch")
    except Exception:
        _tx_rollback(conn, "xlsx_batch")
    if errors and error_out:
        error_out.parent.mkdir(parents=True, exist_ok=True)
        error_out.write_text(json.dumps(errors, ensure_ascii=False, indent=2), encoding="utf-8")
    if errors and not partial:
        e0 = errors[0]
        raise RuntimeError(f"导入失败，共 {len(errors)} 行错误；首个错误：{e0['sheet']} 第{e0['row']}行 {e0['error']}")
    return ok, len(errors)


def show_project_status(conn, project_code: str):
    rows = conn.execute(
        "SELECT * FROM v_project_material_status WHERE project_code=? ORDER BY category, mpn",
        (project_code,),
    ).fetchall()
    if not rows:
        print("没有记录（项目不存在或未建 BOM）")
        return
    # 简洁输出
    headers = ["mpn", "part_desc", "req_qty", "total_stock", "reserved_qty_all_projects",
               "available_stock", "reserved_for_project", "remaining_to_reserve", "shortage_if_reserve_now", "package"]
    print("\t".join(headers))
    for r in rows:
        print("\t".join(str(r[h]) for h in headers))


def show_alloc_detail(conn, project_code: str):
    rows = conn.execute(
        "SELECT * FROM v_project_alloc_detail WHERE project_code=? ORDER BY updated_at DESC",
        (project_code,),
    ).fetchall()
    if not rows:
        print("没有预留记录")
        return
    headers = ["updated_at", "mpn", "location", "alloc_qty", "status", "note"]
    print("\t".join(headers))
    for r in rows:
        print("\t".join(str(r[h]) for h in headers))

def init_locations(conn, room: str, cabinets: list, positions_per_shelf: int = 10, overwrite_note: bool = False):
    """
    cabinets: list of dicts, e.g.
      [{"code":"G01","shelves":3,"note":"三层柜 30x80x35"},
       {"code":"G02","shelves":1,"note":"一层柜 40x100x40"}]
    """
    total = 0
    for cab in cabinets:
        g = cab["code"]
        shelves = int(cab["shelves"])
        cab_note = cab.get("note", "")
        for s in range(1, shelves + 1):
            for p in range(1, positions_per_shelf + 1):
                loc = f"{room}-{g}-S{s:02d}-P{p:02d}"
                note = f"{room} {g} 第{s}层 位{p:02d}"
                if cab_note:
                    note = note + f" | {cab_note}"
                # 插入或忽略（避免重复）
                conn.execute(
                    "INSERT OR IGNORE INTO locations (location, note) VALUES (?, ?)",
                    (loc, note),
                )
                if overwrite_note:
                    conn.execute(
                        "UPDATE locations SET note=? WHERE location=?",
                        (note, loc),
                    )
                total += 1
    return total


def _pick_first(d: dict, names: list[str], default=""):
    for n in names:
        if n in d and clean_text(d.get(n)):
            return clean_text(d.get(n))

    # 兼容 pandas 重名列自动追加后缀（如 Manufacturer.1 / 商品链接.1）
    for key, val in d.items():
        base = re.sub(r"\.\d+$", "", str(key))
        if base in names and clean_text(val):
            return clean_text(val)
    return default


def _to_float(v, default=0.0) -> float:
    s = clean_text(v)
    if not s:
        return default
    s = s.replace(",", "")
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    return float(m.group(0)) if m else default


def _to_int(v, default=0) -> int:
    return int(round(_to_float(v, float(default))))


def _load_lcsc_rows(path: Path) -> list[dict]:
    ext = path.suffix.lower()
    if ext == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            return list(csv.DictReader(f))
    if ext in {".xlsx", ".xls"}:
        try:
            import pandas as pd
        except Exception as e:
            raise RuntimeError("读取 Excel 需要 pandas，请先安装 pandas") from e
        df = pd.read_excel(path)
        # 兼容立创 BOM 报价单：表头通常不在第一行
        first_cols = [str(c) for c in df.columns]
        if not any(x in first_cols for x in ["型号", "Manufacturer Part", "商品名称", "购买数量", "数量"]):
            raw = pd.read_excel(path, header=None)
            header_idx = None
            expected = {"型号", "Manufacturer Part", "商品名称", "购买数量", "数量", "封装", "分类"}
            for i in range(min(len(raw), 30)):
                vals = {clean_text(v) for v in raw.iloc[i].tolist() if clean_text(v)}
                if len(expected.intersection(vals)) >= 3:
                    header_idx = i
                    break
            if header_idx is not None:
                df = pd.read_excel(path, header=header_idx)
        return [
            {str(k): ("" if pd.isna(v) else str(v)) for k, v in row.items()}
            for row in df.to_dict(orient="records")
        ]
    raise RuntimeError("立创导出文件仅支持 .csv/.xlsx/.xls")


def export_project_forms(
    conn,
    project_code: str,
    outbound_csv: Path,
    inbound_csv: Path,
    lcsc_file: Path | None = None,
    inbound_location: str = "",
    apply_inbound: bool = False,
):
    now_str = datetime.now().strftime("%Y-%m-%d")

    # 若提供立创文件：将其视作项目完整 BOM，覆盖 project_bom
    inbound_records = []
    if lcsc_file:
        raw_rows = _load_lcsc_rows(lcsc_file)

        proj_row = conn.execute("SELECT id FROM projects WHERE code=?", (project_code,)).fetchone()
        if proj_row is None:
            conn.execute(
                "INSERT INTO projects (code, name, note) VALUES (?, ?, ?)",
                (project_code, project_code, f"imported_from={lcsc_file.name}"),
            )
            project_id = int(conn.execute("SELECT id FROM projects WHERE code=?", (project_code,)).fetchone()["id"])
        else:
            project_id = int(proj_row["id"])

        bom_qty_by_part: dict[int, int] = {}
        inbound_by_part: dict[int, dict] = {}
        parsed_part_rows = 0

        for row in raw_rows:
            mpn = _pick_first(row, ["型号", "Manufacturer Part", "MPN", "mpn"])
            if not mpn:
                continue
            parsed_part_rows += 1

            name = _pick_first(row, ["商品名称", "Name", "名称"], default=mpn)
            category = _pick_first(row, ["目录", "分类", "Category", "一级分类", "二级分类"], default="立创导入")
            package = _pick_first(row, ["封装", "Footprint 封装", "Footprint", "package"])
            params = _pick_first(row, ["参数", "参数.1", "规格参数", "params"])
            note = _pick_first(row, ["Manufacturer"], default=f"imported_from={lcsc_file.name}")
            url = _pick_first(row, ["商品链接", "商品链接.1"])
            url_norm = normalize_url(url) if url else ""

            part_id = upsert_part(
                conn,
                mpn=mpn,
                name=name or mpn,
                category=category or "立创导入",
                package=package,
                params=params or "",
                url=url_norm,
                datasheet="",
                note=note,
            )

            qty = _to_int(_pick_first(row, ["购买数量", "数量", "Quantity", "qty"], default="0"), 0)
            price = _to_float(_pick_first(row, ["单价(RMB)", "单价", "price"], default="0"), 0.0)

            if qty > 0:
                bom_qty_by_part[part_id] = bom_qty_by_part.get(part_id, 0) + qty

            part_row = conn.execute("SELECT id, mpn, name, unit FROM parts WHERE id=?", (part_id,)).fetchone()
            if part_row is None:
                continue

            rec = inbound_by_part.get(part_id)
            if rec is None:
                rec = {
                    "seq": int(part_row["id"]),
                    "mpn": part_row["mpn"],
                    "part_name": part_row["name"],
                    "unit": part_row["unit"] or "pcs",
                    "qty": 0,
                    "price": 0.0,
                }
                inbound_by_part[part_id] = rec
            rec["qty"] += max(0, qty)
            if price > 0:
                rec["price"] = price

        # 覆盖项目 BOM：xlsx 即该项目完整用料
        # 为避免数量列识别失败导致静默清空 BOM，未解析出任何有效数量时直接报错。
        if parsed_part_rows > 0 and not bom_qty_by_part:
            raise RuntimeError(
                "未从立创文件解析到有效数量（qty>0），已取消覆盖项目 BOM；请检查数量列表头/格式。"
            )
        conn.execute("DELETE FROM project_bom WHERE project_id=?", (project_id,))
        for part_id, req_qty in bom_qty_by_part.items():
            conn.execute(
                "INSERT INTO project_bom (project_id, part_id, req_qty, priority, note) VALUES (?,?,?,?,?)",
                (project_id, part_id, int(req_qty), 2, f"imported_from={lcsc_file.name}"),
            )

        inbound_records = []
        for rec in inbound_by_part.values():
            qty = int(rec["qty"])
            price = float(rec["price"])
            inbound_records.append({
                "seq": rec["seq"],
                "name": rec["mpn"],             # 名称对应 mpn
                "spec": rec["part_name"],       # 型号规格对应 parts.name
                "unit": rec["unit"],
                "qty": qty,
                "price": price,
                "total": qty * price,
                "mpn": rec["mpn"],
            })
        inbound_records.sort(key=lambda x: x["seq"])

    # 出库单：基于（可能刚覆盖后的）项目 BOM
    out_rows = conn.execute(
        """
        SELECT p.id, p.mpn, p.name, p.unit, b.req_qty
        FROM project_bom b
        JOIN projects pr ON pr.id = b.project_id
        JOIN parts p ON p.id = b.part_id
        WHERE pr.code = ?
        ORDER BY p.category, p.mpn
        """,
        (project_code,),
    ).fetchall()
    outbound_records = [
        {
            "seq": int(r["id"]),
            "name": r["mpn"],
            "spec": r["name"],
            "unit": r["unit"] or "pcs",
            "qty": int(r["req_qty"]),
            "price": 0.0,
            "total": 0.0,
            "mpn": r["mpn"],
        }
        for r in out_rows
    ]

    # 未提供立创文件时，入库单回退为项目 BOM 需求数量
    if not lcsc_file:
        inbound_records = list(outbound_records)

    if not outbound_records:
        raise RuntimeError(f"项目未找到或 BOM 为空：{project_code}")

    outbound_csv.parent.mkdir(parents=True, exist_ok=True)
    with outbound_csv.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["序号", "时间", "名称", "型号规格", "单位", "数量", "单价(元)", "总额(元)", "项目"])
        for r in outbound_records:
            w.writerow([
                r["seq"],
                now_str,
                r["name"],
                r["spec"],
                r["unit"],
                r["qty"],
                f"{r['price']:.4f}" if r["price"] else "",
                f"{r['total']:.4f}" if r["total"] else "",
                project_code,
            ])

    inbound_csv.parent.mkdir(parents=True, exist_ok=True)
    with inbound_csv.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["序号", "时间", "名称", "型号规格", "单位", "数量", "单价(元)", "总额(元)", "项目"])
        for r in inbound_records:
            w.writerow([
                r["seq"],
                now_str,
                r["name"],
                r["spec"],
                r["unit"],
                r["qty"],
                f"{r['price']:.4f}" if r["price"] else "",
                f"{r['total']:.4f}" if r["total"] else "",
                project_code,
            ])

    if apply_inbound:
        if not inbound_location:
            raise RuntimeError("执行入库写库时必须提供 --inbound-loc")
        for r in inbound_records:
            if int(r["qty"]) <= 0:
                continue
            row = conn.execute("SELECT id FROM parts WHERE mpn=?", (r["mpn"],)).fetchone()
            if row is None:
                upsert_part(
                    conn,
                    mpn=r["mpn"],
                    name=r["spec"] or r["mpn"],
                    category="立创导入",
                    package="",
                    params="",
                    url="",
                    datasheet="",
                    note=f"project={project_code}",
                )
            add_stock(conn, r["mpn"], inbound_location, int(r["qty"]), "new", f"project={project_code} lcsc入库")


def import_lcsc_file_to_parts_and_stock(
    conn,
    lcsc_file: Path,
    inbound_location: str = "LCSC-INBOX",
    datasheets_dir: Path | None = None,
):
    raw_rows = _load_lcsc_rows(lcsc_file)
    part_written = 0
    stock_written = 0

    conn.execute(
        "INSERT OR IGNORE INTO locations (location, note) VALUES (?, ?)",
        (inbound_location, "自动创建：立创导入默认入库位"),
    )

    datasheets_dir = datasheets_dir or (Path(conn.execute("PRAGMA database_list").fetchone()[2]).parent / "datasheets")

    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36",
        "Accept-Language": "zh-CN,zh;q=0.9",
    })

    for row in raw_rows:
        mpn = _pick_first(row, ["型号", "Manufacturer Part", "MPN", "mpn"])
        name = _pick_first(row, ["商品名称", "Name", "名称"], default=mpn)
        if not mpn:
            continue

        qty = _to_int(_pick_first(row, ["购买数量", "数量", "Quantity", "qty"], default="0"), 0)
        package = _pick_first(row, ["封装", "Footprint 封装", "Footprint", "package"])
        category = _pick_first(row, ["目录", "分类", "Category", "一级分类", "二级分类"], default="立创导入")
        params = _pick_first(row, ["参数", "参数.1", "规格参数", "params"])
        note = _pick_first(row, ["Manufacturer"], default=f"imported_from={lcsc_file.name}")
        supplier_part = _pick_first(row, ["商品编号", "Supplier Part", "LCSC", "LCSC编号"])
        url = _pick_first(row, ["商品链接", "商品链接.1"])
        url_norm = normalize_url(url) if url else ""

        datasheet_local = ""
        if url_norm:
            try:
                page_resp = session.get(url_norm, timeout=20)
                page_resp.raise_for_status()
                page_resp.encoding = page_resp.apparent_encoding or "utf-8"
                soup = BeautifulSoup(page_resp.text, "lxml")
                pdf_url = find_datasheet_url(soup, url_norm)
                if pdf_url:
                    base = safe_filename(mpn)
                    if supplier_part:
                        base += f"__{safe_filename(supplier_part)}"
                    out_pdf = datasheets_dir / f"{base}.pdf"
                    if download_pdf(session, pdf_url, out_pdf):
                        datasheet_local = str(out_pdf)
            except Exception:
                pass

        part_id = upsert_part(
            conn,
            mpn=mpn,
            name=name or mpn,
            category=category or "立创导入",
            package=package,
            params=params or "",
            url=url_norm,
            datasheet=datasheet_local,
            note=note,
        )
        if part_id:
            part_written += 1

        if qty > 0:
            add_stock(conn, mpn, inbound_location, qty, "new", f"imported_from={lcsc_file.name}")
            stock_written += 1

    return part_written, stock_written


# ---------------------------
# CLI
# ---------------------------
def main():
    ap = argparse.ArgumentParser(description="实验室元器件/设备库存 + 项目管理（SQLite）")
    ap.add_argument("--db", required=True, help=r"数据库路径，例如 G:\LabInventory\lab_inventory.db")

    sub = ap.add_subparsers(dest="cmd", required=True)
    
    # init locations
    p = sub.add_parser("init-locations", help="初始化库位编码到 locations 表（默认按柜子/层/位生成）")
    p.add_argument("--room", default="C409", help="房间号，例如 C409")
    p.add_argument("--g01-shelves", type=int, default=3, help="G01 层数（默认3）")
    p.add_argument("--g02-shelves", type=int, default=1, help="G02 层数（默认1）")
    p.add_argument("--positions", type=int, default=10, help="每层位置数（默认10：P01~P10）")
    p.add_argument("--overwrite-note", action="store_true", help="覆盖更新 locations.note（默认不覆盖）")

    # lcsc import
    p = sub.add_parser("lcsc", help="从立创商品链接导入物料并自动下载数据手册")
    p.add_argument("--url", required=True, help="立创商品链接")
    p.add_argument("--datasheets-dir", default="", help=r"数据手册保存目录（默认=数据库同级 datasheets）")

    # stock in
    p = sub.add_parser("stock-in", help="入库（按库位增加库存）")
    p.add_argument("--mpn", required=True)
    p.add_argument("--loc", required=True)
    p.add_argument("--qty", type=int, required=True)
    p.add_argument("--condition", default="new")
    p.add_argument("--note", default="")

    p = sub.add_parser("stock-out", help="出库（按库位扣减库存）")
    p.add_argument("--mpn", required=True)
    p.add_argument("--loc", required=True)
    p.add_argument("--qty", type=int, required=True)
    p.add_argument("--proj", default="", help="可选项目 code")
    p.add_argument("--ref", default="")
    p.add_argument("--note", default="")
    p.add_argument("--operator", default="")

    p = sub.add_parser("stock-move", help="移库（from 扣减 + to 增加）")
    p.add_argument("--mpn", required=True)
    p.add_argument("--from", dest="from_loc", required=True)
    p.add_argument("--to", dest="to_loc", required=True)
    p.add_argument("--qty", type=int, required=True)
    p.add_argument("--note", default="")
    p.add_argument("--operator", default="")

    p = sub.add_parser("stock-adjust", help="库存调整（--add 或 --sub 二选一，需说明原因）")
    p.add_argument("--mpn", required=True)
    p.add_argument("--loc", required=True)
    g = p.add_mutually_exclusive_group(required=True)
    g.add_argument("--add", type=int, default=0)
    g.add_argument("--sub", type=int, default=0)
    p.add_argument("--note", required=True)
    p.add_argument("--ref", default="")
    p.add_argument("--operator", default="")

    # project create
    p_project = sub.add_parser("project", help="项目管理（新增资源挂接）")
    project_sub = p_project.add_subparsers(dest="project_cmd", required=True)

    p = project_sub.add_parser("add", help="创建项目（存在则更新）")
    p.add_argument("--code", required=True)
    p.add_argument("--name", required=True)
    p.add_argument("--owner", default="")
    p.add_argument("--note", default="")

    p = project_sub.add_parser("overview", help="查看项目总览")
    p.add_argument("--code", default="", help="可选，项目 code")

    p_res = project_sub.add_parser("resource", help="项目资源管理")
    resource_sub = p_res.add_subparsers(dest="resource_cmd", required=True)

    p = resource_sub.add_parser("add", help="添加/更新项目资源")
    p.add_argument("--code", required=True)
    p.add_argument("--type", required=True)
    p.add_argument("--name", required=True)
    p.add_argument("--uri", required=True)
    p.add_argument("--is-dir", type=int, choices=[0, 1], default=1)
    p.add_argument("--tags", default="")
    p.add_argument("--note", default="")
    p.add_argument("--no-check", action="store_true")

    p = resource_sub.add_parser("ls", help="列出项目资源")
    p.add_argument("--code", required=True)

    p = resource_sub.add_parser("rm", help="删除项目资源")
    p.add_argument("--code", required=True)
    p.add_argument("--type", required=True)
    p.add_argument("--uri", required=True)

    p = resource_sub.add_parser("check", help="检查项目资源有效性")
    p.add_argument("--code", required=True)

    p = resource_sub.add_parser("import-xlsx", help="批量导入项目资源")
    p.add_argument("--xlsx", required=True)
    p.add_argument("--sheet", default="project_resources")
    p.add_argument("--header-row", type=int, default=1)
    p.add_argument("--auto-create-project", action="store_true")
    p.add_argument("--no-check", action="store_true")

    p = sub.add_parser("proj-new", help="创建项目")
    p.add_argument("--code", required=True)
    p.add_argument("--name", required=True)
    p.add_argument("--owner", default="")
    p.add_argument("--note", default="")

    # bom set
    p = sub.add_parser("bom-set", help="设置/更新项目BOM需求")
    p.add_argument("--proj", required=True, help="项目 code，例如 PJ-001")
    p.add_argument("--mpn", required=True)
    p.add_argument("--req", type=int, required=True, help="需求数量")
    p.add_argument("--priority", type=int, default=2)
    p.add_argument("--note", default="")

    # reserve
    p = sub.add_parser("reserve", help="带库位预留（强约束：禁止超预留）")
    p.add_argument("--proj", required=True)
    p.add_argument("--mpn", required=True)
    p.add_argument("--loc", required=True)
    p.add_argument("--qty", type=int, required=True)
    p.add_argument("--note", default="")

    # release
    p = sub.add_parser("release", help="释放预留（按 alloc_id）")
    p.add_argument("--id", type=int, required=True)
    p.add_argument("--note", default="释放")

    # consume
    p = sub.add_parser("consume", help="消耗预留（按 alloc_id），同时扣减 stock")
    p.add_argument("--id", type=int, required=True)
    p.add_argument("--note", default="已消耗")

    # show status
    p = sub.add_parser("proj-status", help="查看项目备料状态（BOM+库存+预留）")
    p.add_argument("--proj", required=True)

    # show alloc
    p = sub.add_parser("proj-alloc", help="查看项目预留明细（带库位）")
    p.add_argument("--proj", required=True)

    p = sub.add_parser("ledger", help="查询库存流水")
    p.add_argument("--proj", default="", help="按项目 code 过滤")
    p.add_argument("--mpn", default="", help="按物料 MPN 过滤")
    p.add_argument("--since", default="", help="按日期过滤（YYYY-MM-DD）")

    p = sub.add_parser("schema-export", help="导出数据库结构（tables/indexes/views/triggers）")
    p.add_argument("--format", dest="fmt", choices=["sql", "md"], default="sql")
    p.add_argument("--out", default="", help="输出文件路径；不提供则输出到 stdout")

    p = sub.add_parser("txn-export-xlsx", help="导出交易模板 xlsx")
    p.add_argument("--out", required=True, help="输出 xlsx 路径")

    p = sub.add_parser("txn-import-xlsx", help="导入交易 xlsx（Transactions sheet）")
    p.add_argument("--xlsx", required=True, help="输入 xlsx 文件")
    p.add_argument("--mode", choices=["auto", "transactions", "stock-io"], default="auto", help="导入模式：auto(自动识别) / transactions / stock-io")
    p.add_argument("--partial", action="store_true", help="开启部分成功模式（默认全有全无）")
    p.add_argument("--error-out", default="", help="错误报告输出 json 路径")

    # project forms
    p = sub.add_parser("proj-forms", help="按项目生成出库/入库单 CSV；或仅按立创文件自动写入 parts+stock")
    p.add_argument("--proj", default="", help="项目 code")
    p.add_argument("--outbound-csv", default="", help="导出的出库单 CSV 路径")
    p.add_argument("--inbound-csv", default="", help="导出的入库单 CSV 路径")
    p.add_argument("--lcsc-file", default="", help="立创导出文件（csv/xlsx/xls），用于填充入库单")
    p.add_argument("--apply-inbound", action="store_true", help="将入库单数量写入库存")
    p.add_argument("--inbound-loc", default="", help="执行 --apply-inbound 时的入库库位")

    args = ap.parse_args()
    cwd = Path.cwd()
    db_path = resolve_input_path(args.db, cwd)

    if not db_path.exists():
        raise SystemExit(f"数据库不存在：{db_path}")

    conn = connect(db_path)
    try:
        init_db(conn)  # 幂等补齐结构/触发器/视图
        if args.cmd == "init-locations":
            cabinets = [
                {"code": "G01", "shelves": args.g01_shelves, "note": "三层柜 30cm深x80cm长x35cm高"},
                {"code": "G02", "shelves": args.g02_shelves, "note": "一层柜 40cm深x100cm长x40cm高"},
            ]
            n = init_locations(conn, room=args.room, cabinets=cabinets, positions_per_shelf=args.positions, overwrite_note=args.overwrite_note)
            conn.commit()
            print(f"库位初始化完成：写入/确保存在 {n} 个库位（{args.room}）")
            return
        
        if args.cmd == "lcsc":
            datasheets_dir = Path(args.datasheets_dir) if args.datasheets_dir else (db_path.parent / "datasheets")
            item = lcsc_fetch_and_parse(args.url, datasheets_dir)
            part_id = upsert_part(
                conn,
                mpn=item.mpn,
                name=item.desc,              # 你要求：一句话描述
                category=item.category,
                package=item.package,
                params=item.params_text,      # 参数表进 params
                url=item.page_url,
                datasheet=item.datasheet_local or item.page_url,
                note=item.note,
            )
            conn.commit()
            print(f"导入完成：mpn={item.mpn} part_id={part_id}")
            print(f"datasheet={item.datasheet_local or item.page_url}")
            return

        if args.cmd == "stock-in":
            stock_in(conn, args.mpn, args.loc, args.qty, args.condition, args.note)
            conn.commit()
            print(f"入库成功：{args.mpn} @ {args.loc} +{args.qty}")
            return

        if args.cmd == "stock-out":
            stock_out(conn, args.mpn, args.loc, args.qty, args.proj, args.ref, args.note, args.operator)
            conn.commit()
            print(f"出库成功：{args.mpn} @ {args.loc} -{args.qty}")
            return

        if args.cmd == "stock-move":
            stock_move(conn, args.mpn, args.from_loc, args.to_loc, args.qty, args.note, args.operator)
            conn.commit()
            print(f"移库成功：{args.mpn} {args.from_loc} -> {args.to_loc} qty={args.qty}")
            return

        if args.cmd == "stock-adjust":
            stock_adjust(conn, args.mpn, args.loc, add_qty=args.add, sub_qty=args.sub, note=args.note, ref=args.ref, operator=args.operator)
            conn.commit()
            mode = "add" if args.add > 0 else "sub"
            v = args.add if args.add > 0 else args.sub
            print(f"调整成功：{args.mpn} @ {args.loc} {mode} {v}")
            return

        if args.cmd == "project":
            if args.project_cmd == "add":
                _, created = add_project(conn, args.code, args.name, args.owner, args.note)
                conn.commit()
                print(f"项目{'创建' if created else '更新'}成功：{args.code}")
                return
            if args.project_cmd == "overview":
                show_project_overview(conn, args.code)
                return
            if args.project_cmd == "resource":
                if args.resource_cmd == "add":
                    project_id = get_project_id(conn, args.code)
                    rid = upsert_resource(
                        conn,
                        project_id=project_id,
                        resource_type=args.type,
                        name=args.name,
                        uri=args.uri,
                        is_dir=args.is_dir,
                        tags=args.tags,
                        note=args.note,
                        no_check=args.no_check,
                    )
                    conn.commit()
                    print(f"项目资源写入成功：id={rid}")
                    return
                if args.resource_cmd == "ls":
                    show_project_resources(conn, args.code)
                    return
                if args.resource_cmd == "rm":
                    project_id = get_project_id(conn, args.code)
                    n = remove_resource(conn, project_id, args.type, args.uri)
                    conn.commit()
                    print(f"删除完成：{n} 条")
                    return
                if args.resource_cmd == "check":
                    check_project_resources(conn, args.code)
                    return
                if args.resource_cmd == "import-xlsx":
                    xlsx_path = resolve_input_path(args.xlsx, cwd)
                    ok, err = import_resources_xlsx(
                        conn,
                        xlsx_path=xlsx_path,
                        sheet=args.sheet,
                        header_row=args.header_row,
                        no_check=args.no_check,
                        auto_create_project=args.auto_create_project,
                        get_project_id=lambda code: get_project_id(conn, code),
                        create_project=lambda code, name: add_project(conn, code, name),
                    )
                    conn.commit()
                    print(f"导入完成：成功 {ok} 行，失败 {err} 行")
                    return

        if args.cmd == "proj-new":
            create_project(conn, args.code, args.name, args.owner, args.note)
            conn.commit()
            print(f"项目创建成功：{args.code} {args.name}")
            return

        if args.cmd == "bom-set":
            set_bom(conn, args.proj, args.mpn, args.req, args.priority, args.note)
            conn.commit()
            print(f"BOM已更新：{args.proj} {args.mpn} req={args.req}")
            return

        if args.cmd == "reserve":
            alloc_id = reserve_loc(conn, args.proj, args.mpn, args.loc, args.qty, args.note)
            conn.commit()
            print(f"预留成功：alloc_id={alloc_id} {args.proj} {args.mpn} {args.loc} +{args.qty}")
            return

        if args.cmd == "release":
            release_alloc(conn, args.id, args.note)
            conn.commit()
            print(f"释放成功：alloc_id={args.id}")
            return

        if args.cmd == "consume":
            consume_alloc(conn, args.id, args.note)
            conn.commit()
            print(f"消耗成功：alloc_id={args.id}")
            return

        if args.cmd == "proj-status":
            show_project_status(conn, args.proj)
            return

        if args.cmd == "proj-alloc":
            show_alloc_detail(conn, args.proj)
            return

        if args.cmd == "ledger":
            show_ledger(conn, args.proj, args.mpn, args.since)
            return

        if args.cmd == "schema-export":
            out_path = resolve_output_path(args.out, cwd) if args.out else None
            schema_export(conn, fmt=args.fmt, out_path=out_path)
            if out_path:
                print(f"schema 已导出：{out_path}")
            return

        if args.cmd == "txn-export-xlsx":
            out = resolve_output_path(args.out, cwd)
            txn_export_xlsx_template(out)
            print(f"交易模板已导出：{out}")
            return

        if args.cmd == "txn-import-xlsx":
            xlsx_path = resolve_input_path(args.xlsx, cwd)
            error_out = resolve_output_path(args.error_out, cwd) if args.error_out else None
            ok, err = txn_import_xlsx(conn, xlsx_path, partial=args.partial, error_out=error_out, mode=args.mode)
            conn.commit()
            print(f"导入完成：成功 {ok} 行，失败 {err} 行")
            if error_out:
                print(f"错误报告：{error_out}")
            return

        if args.cmd == "proj-forms":
            if args.lcsc_file and not args.proj and not args.outbound_csv and not args.inbound_csv:
                loc = args.inbound_loc or "LCSC-INBOX"
                part_written, stock_written = import_lcsc_file_to_parts_and_stock(
                    conn,
                    lcsc_file=resolve_input_path(args.lcsc_file, cwd),
                    inbound_location=loc,
                    datasheets_dir=(db_path.parent / "datasheets"),
                )
                conn.commit()
                print(f"立创文件已导入：{args.lcsc_file}")
                print(f"parts 写入/更新：{part_written} 条")
                print(f"stock 写入：{stock_written} 条，库位={loc}")
                return

            if not (args.proj and args.outbound_csv and args.inbound_csv):
                raise RuntimeError(
                    "proj-forms 生成单据模式需要同时提供 --proj --outbound-csv --inbound-csv；"
                    "若仅想导入立创文件到 parts/stock，可只提供 --lcsc-file"
                )

            export_project_forms(
                conn,
                project_code=args.proj,
                outbound_csv=resolve_output_path(args.outbound_csv, cwd),
                inbound_csv=resolve_output_path(args.inbound_csv, cwd),
                lcsc_file=resolve_input_path(args.lcsc_file, cwd) if args.lcsc_file else None,
                inbound_location=args.inbound_loc,
                apply_inbound=args.apply_inbound,
            )
            if args.apply_inbound or args.lcsc_file:
                conn.commit()
            print(f"出库单已生成：{args.outbound_csv}")
            print(f"入库单已生成：{args.inbound_csv}")
            if args.apply_inbound:
                print(f"已按入库单写库：loc={args.inbound_loc}")
            return

    except sqlite3.IntegrityError as e:
        # 触发器/约束报错通常在这里
        conn.rollback()
        raise SystemExit(f"数据库约束失败：{e}")
    except Exception as e:
        conn.rollback()
        raise SystemExit(f"执行失败：{e}")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
